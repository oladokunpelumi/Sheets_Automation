"""
Twitter/X Verified-Only Scanner (Lightweight)
Optimized for larger batches (1000+ usernames).
Only checks if a username has a blue check — skips follower extraction.

Usage:
  python verified_only_scanner.py input.xlsx --sheet ">1000" --output verified_results.xlsx
  python verified_only_scanner.py input.xlsx --sheet ">1000" --output verified_results.xlsx --headless
"""

import argparse
import json
import os
import re
import socket
import sys
import time
import random
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout


# ── Configuration ────────────────────────────────────────────────────────────

MIN_DELAY = 2               # Slightly faster since we do less per page
MAX_DELAY = 4
BATCH_PAUSE_EVERY = 50
BATCH_PAUSE_SECONDS = 45    # Longer pause for large batches
PAGE_TIMEOUT = 15000         # Shorter timeout — we only need partial load
CHECKPOINT_EVERY = 25
MAX_RETRIES = 2


# ── Helper: Wait for internet connection ─────────────────────────────────────

def wait_for_internet():
    """Blocks and waits until an active internet connection is detected."""
    def is_connected():
        try:
            # Connect to Cloudflare DNS to check internet
            socket.create_connection(("1.1.1.1", 53), timeout=3)
            return True
        except OSError:
            pass
        return False

    if not is_connected():
        print("\n[!] Internet disconnected! Pausing scan... waiting for connection to resume.", flush=True)
        while not is_connected():
            time.sleep(5)
        print("[✓] Internet connection restored! Resuming scan...\n", flush=True)
        time.sleep(3)

# ── Core: Check verified status only ────────────────────────────────────────

def check_verified(page, username: str) -> dict:
    url = f"https://x.com/{username}"
    result = {"username": username, "verified": "No", "status": "OK"}

    try:
        page.goto(url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT)
        page.wait_for_timeout(2500)

        page_text = page.content()

        if "This account doesn't exist" in page_text or "Account suspended" in page_text:
            result["status"] = "Not Found / Suspended"
            return result

        if "These tweets are protected" in page_text:
            result["status"] = "Private Account"

        # Check for verified badge via selectors
        verified_selectors = [
            '[data-testid="icon-verified"]',
            'svg[aria-label="Verified account"]',
            'svg[aria-label="Verified"]',
            '[aria-label="Provides a verified blue checkmark"]',
            'a[href="/i/verified-choose"] svg',
        ]

        for sel in verified_selectors:
            try:
                if page.locator(sel).first.is_visible(timeout=800):
                    result["verified"] = "Yes"
                    return result
            except Exception:
                continue

        # Fallback: raw HTML check
        if ('aria-label="Verified"' in page_text
            or 'data-testid="icon-verified"' in page_text
            or "Verified account" in page_text):
            result["verified"] = "Yes"

    except PlaywrightTimeout:
        result["status"] = "Timeout"
    except Exception as e:
        result["status"] = f"Error: {str(e)[:60]}"

    return result


# ── Read usernames from xlsx ─────────────────────────────────────────────────

def read_usernames(filepath: str, sheet_name: str = None, column: str = "A") -> list[str]:
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print(f"Error: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
            sys.exit(1)
        ws = wb[sheet_name]
    else:
        ws = wb.active

    usernames = []
    col_idx = openpyxl.utils.column_index_from_string(column)
    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
        val = row[0]
        if val and str(val).strip():
            usernames.append(str(val).strip().lstrip("@"))

    wb.close()
    print(f"Loaded {len(usernames)} usernames from '{filepath}' (sheet: {sheet_name or 'active'})")
    return usernames


# ── Write results to xlsx ────────────────────────────────────────────────────

def write_results(results: list[dict], output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Verified Check Results"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1DA1F2")
    header_align = Alignment(horizontal="center", vertical="center")
    data_font = Font(name="Arial", size=10)
    data_align_center = Alignment(horizontal="center", vertical="center")
    data_align_left = Alignment(horizontal="left", vertical="center")
    yes_fill = PatternFill("solid", fgColor="D4EDDA")
    no_fill = PatternFill("solid", fgColor="F8D7DA")
    error_fill = PatternFill("solid", fgColor="FFF3CD")
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    headers = ["#", "Username", "Blue Check Verified (Yes/No)", "Status"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for i, r in enumerate(results, start=1):
        row_num = i + 1
        is_error = r["status"] not in ("OK", "Private Account")

        ws.cell(row=row_num, column=1, value=i).font = data_font
        ws.cell(row=row_num, column=1).alignment = data_align_center
        ws.cell(row=row_num, column=1).border = thin_border

        cell_user = ws.cell(row=row_num, column=2, value=f'@{r["username"]}')
        cell_user.font = Font(name="Arial", size=10, color="1DA1F2")
        cell_user.alignment = data_align_left
        cell_user.border = thin_border

        cell_v = ws.cell(row=row_num, column=3, value=r["verified"])
        cell_v.font = data_font
        cell_v.alignment = data_align_center
        cell_v.border = thin_border
        if r["verified"] == "Yes":
            cell_v.fill = yes_fill
        elif not is_error:
            cell_v.fill = no_fill

        cell_s = ws.cell(row=row_num, column=4, value=r["status"])
        cell_s.font = data_font
        cell_s.alignment = data_align_center
        cell_s.border = thin_border
        if is_error:
            cell_s.fill = error_fill

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{len(results) + 1}"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    total = len(results)
    verified_count = sum(1 for r in results if r["verified"] == "Yes")
    not_verified = sum(1 for r in results if r["verified"] == "No" and r["status"] == "OK")
    errors = sum(1 for r in results if r["status"] not in ("OK", "Private Account"))

    summary_data = [
        ("Metric", "Value"),
        ("Total Usernames Scanned", total),
        ("Blue Check Verified (Yes)", verified_count),
        ("Not Verified (No)", not_verified),
        ("Errors / Not Found", errors),
        ("Verification Rate", f"{verified_count / max(total, 1) * 100:.1f}%"),
    ]

    for row_idx, (label, value) in enumerate(summary_data, 1):
        cell_a = ws2.cell(row=row_idx, column=1, value=label)
        cell_b = ws2.cell(row=row_idx, column=2, value=value)
        cell_a.border = thin_border
        cell_b.border = thin_border
        if row_idx == 1:
            cell_a.font = header_font
            cell_a.fill = header_fill
            cell_b.font = header_font
            cell_b.fill = header_fill
        else:
            cell_a.font = Font(name="Arial", size=10, bold=True)
            cell_b.font = data_font
            cell_b.alignment = data_align_center

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 18

    wb.save(output_path)
    print(f"\nResults saved to: {output_path}")
    print(f"  Total: {total} | Verified: {verified_count} | Not Verified: {not_verified} | Errors: {errors}")


# ── Checkpoint helpers ───────────────────────────────────────────────────────

def save_checkpoint(results, path):
    with open(path, "w") as f:
        json.dump(results, f, indent=2)

def load_checkpoint(path):
    if os.path.exists(path):
        with open(path, "r") as f:
            data = json.load(f)
        print(f"Resuming from checkpoint: {len(data)} already scanned.")
        return data
    return []


# ── Main scan loop ───────────────────────────────────────────────────────────

def run_scan(input_file, sheet_name=None, column="A", output_file="verified_results.xlsx", headless=False, resume=True):
    usernames = read_usernames(input_file, sheet_name, column)
    if not usernames:
        print("No usernames found.")
        return

    checkpoint_path = output_file.replace(".xlsx", "_checkpoint.json")
    results = load_checkpoint(checkpoint_path) if resume else []
    already_done = {r["username"].lower() for r in results}
    remaining = [u for u in usernames if u.lower() not in already_done]

    if not remaining:
        print("All done. Writing output...")
        write_results(results, output_file)
        return

    print(f"\nRemaining: {len(remaining)} | Mode: {'Headless' if headless else 'Visible'}")
    print(f"Delay: {MIN_DELAY}-{MAX_DELAY}s | Batch pause: {BATCH_PAUSE_SECONDS}s every {BATCH_PAUSE_EVERY}\n")

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=headless,
            args=["--disable-blink-features=AutomationControlled"],
        )
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
            viewport={"width": 1280, "height": 900},
            locale="en-US",
        )
        page = context.new_page()

        # Block images, media, fonts — we only need the HTML
        page.route("**/*", lambda route: route.abort() if route.request.resource_type in ("image", "media", "font", "stylesheet") else route.continue_())

        scan_count = 0
        for i, username in enumerate(remaining):
            scan_count += 1
            progress = len(results) + 1
            total = len(usernames)
            pct = progress / total * 100

            print(f"  [{progress}/{total}] ({pct:.0f}%) @{username}...", end=" ", flush=True)

            result = None
            for attempt in range(1, MAX_RETRIES + 1):
                wait_for_internet()
                result = check_verified(page, username)
                if result["status"] in ("OK", "Private Account", "Not Found / Suspended"):
                    break
                
                # Check for explicit network loss
                if "net::ERR_INTERNET_DISCONNECTED" in result["status"] or result["status"] == "Timeout" or "Error" in result["status"]:
                    wait_for_internet()
                    
                if attempt < MAX_RETRIES:
                    print(f"(retry {attempt})...", end=" ", flush=True)
                    time.sleep(2)

            label = "✓ VERIFIED" if result["verified"] == "Yes" else "✗ No"
            print(f"{label} | {result['status']}")
            results.append(result)

            if scan_count % CHECKPOINT_EVERY == 0:
                save_checkpoint(results, checkpoint_path)
                print(f"    ↳ Checkpoint saved ({len(results)} scanned)")

            if scan_count % BATCH_PAUSE_EVERY == 0 and i < len(remaining) - 1:
                print(f"\n  ⏸  Batch pause ({BATCH_PAUSE_SECONDS}s)...\n")
                time.sleep(BATCH_PAUSE_SECONDS)
            else:
                time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))

        browser.close()

    save_checkpoint(results, checkpoint_path)
    write_results(results, output_file)
    if os.path.exists(checkpoint_path):
        os.remove(checkpoint_path)


def main():
    parser = argparse.ArgumentParser(description="Verified-only Twitter/X scanner (lightweight)")
    parser.add_argument("input", help="Path to .xlsx file")
    parser.add_argument("--sheet", default=None, help="Sheet name")
    parser.add_argument("--column", default="A", help="Column with usernames (default: A)")
    parser.add_argument("--output", default="verified_results.xlsx", help="Output file")
    parser.add_argument("--headless", action="store_true", help="Run headless")
    parser.add_argument("--no-resume", action="store_true", help="Ignore checkpoint")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"Error: '{args.input}' not found.")
        sys.exit(1)

    run_scan(args.input, args.sheet, args.column, args.output, args.headless, not args.no_resume)


if __name__ == "__main__":
    main()
