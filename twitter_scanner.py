"""
Twitter/X Profile Scanner
Scans a list of usernames from an .xlsx file and checks:
  - Blue check (verified) status
  - Follower count
Outputs results to a new professionally formatted .xlsx file.

Usage:
  python twitter_scanner.py input.xlsx --sheet "Sheet1" --output results.xlsx
  python twitter_scanner.py input.xlsx --sheet "Sheet1" --output results.xlsx --column A --headless
"""

import argparse
import json
import os
import re
import socket
import sys
import time
import random
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout


# ── Configuration ────────────────────────────────────────────────────────────

# Delays between requests (seconds) — randomized within this range
MIN_DELAY = 3
MAX_DELAY = 6

# Extra delay every N requests to avoid detection
BATCH_PAUSE_EVERY = 50
BATCH_PAUSE_SECONDS = 30

# Playwright timeout for page load (ms)
PAGE_TIMEOUT = 20000

# Checkpoint: save progress every N usernames
CHECKPOINT_EVERY = 25

# Retry failed lookups up to this many times
MAX_RETRIES = 2


# ── Helper: Wait for internet connection ─────────────────────────────────────

def wait_for_internet():
    """Blocks and waits until an active internet connection is detected."""
    def is_connected():
        try:
            # Connect to Cloudflare DNS to check internet
            socket.create_connection(("1.1.1.1", 53), timeout=3)
            return True
        except OSError:
            pass
        return False

    if not is_connected():
        print("\n[!] Internet disconnected! Pausing scan... waiting for connection to resume.", flush=True)
        while not is_connected():
            time.sleep(5)
        print("[✓] Internet connection restored! Resuming scan...\n", flush=True)
        time.sleep(3)  # Give DNS a moment to stabilize

# ── Helper: Parse follower count strings ─────────────────────────────────────

def parse_follower_count(text: str) -> int | None:
    """Parse '1,234 Followers', '12.5K Followers', '1.2M Followers' etc."""
    if not text:
        return None
    text = text.strip().split()[0].replace(",", "")
    multiplier = 1
    if text.upper().endswith("K"):
        multiplier = 1_000
        text = text[:-1]
    elif text.upper().endswith("M"):
        multiplier = 1_000_000
        text = text[:-1]
    elif text.upper().endswith("B"):
        multiplier = 1_000_000_000
        text = text[:-1]
    try:
        return int(float(text) * multiplier)
    except ValueError:
        return None


# ── Helper: Format follower count for display ────────────────────────────────

def format_followers(count: int | None) -> str:
    if count is None:
        return "N/A"
    if count >= 1_000_000:
        return f"{count / 1_000_000:.1f}M"
    if count >= 1_000:
        return f"{count / 1_000:.1f}K"
    return str(count)


# ── Core: Scrape a single Twitter/X profile ──────────────────────────────────

def scrape_profile(page, username: str) -> dict:
    """
    Visit x.com/<username> and extract verified status + follower count.
    Returns dict with keys: username, verified, followers, followers_raw, status
    """
    url = f"https://x.com/{username}"
    result = {
        "username": username,
        "verified": "No",
        "followers": "N/A",
        "followers_raw": None,
        "status": "OK",
    }

    try:
        response = page.goto(url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT)

        # Wait for profile content to appear
        page.wait_for_timeout(3000)

        # ── Check if account exists / is suspended ───────────────────────
        page_text = page.content()

        if "This account doesn't exist" in page_text or "Account suspended" in page_text:
            result["status"] = "Not Found / Suspended"
            return result

        if "These tweets are protected" in page_text:
            result["status"] = "Private Account"

        # ── Check for verified badge (blue checkmark) ────────────────────
        # Twitter/X uses an SVG icon near the username for verified accounts.
        # We look for multiple possible selectors.
        verified_selectors = [
            # The verified badge SVG typically has this aria-label or data-testid
            '[data-testid="icon-verified"]',
            'svg[aria-label="Verified account"]',
            'svg[aria-label="Verified"]',
            # Sometimes it's inside a specific container
            '[aria-label="Provides a verified blue checkmark"]',
            # Blue check badge container
            'a[href="/i/verified-choose"] svg',
        ]

        for sel in verified_selectors:
            try:
                if page.locator(sel).first.is_visible(timeout=1000):
                    result["verified"] = "Yes"
                    break
            except Exception:
                continue

        # Fallback: search raw HTML for verified indicators
        if result["verified"] == "No":
            # Look for common verified badge patterns in the page source
            if ('aria-label="Verified"' in page_text
                or 'data-testid="icon-verified"' in page_text
                or "Verified account" in page_text):
                result["verified"] = "Yes"

        # ── Extract follower count ───────────────────────────────────────
        # Follower count is typically in a link like: /username/followers
        followers_selectors = [
            f'a[href="/{username}/verified_followers"] span span',
            f'a[href="/{username}/followers"] span span',
            f'a[href="/{username}/followers"] span',
        ]

        for sel in followers_selectors:
            try:
                elements = page.locator(sel).all()
                for el in elements:
                    text = el.inner_text(timeout=2000)
                    if text and any(c.isdigit() for c in text):
                        count = parse_follower_count(text)
                        if count is not None:
                            result["followers_raw"] = count
                            result["followers"] = format_followers(count)
                            break
                if result["followers_raw"] is not None:
                    break
            except Exception:
                continue

        # Fallback: regex search in page text for follower patterns
        if result["followers_raw"] is None:
            match = re.search(r'([\d,.]+[KMB]?)\s*Followers', page_text)
            if match:
                count = parse_follower_count(match.group(1))
                if count is not None:
                    result["followers_raw"] = count
                    result["followers"] = format_followers(count)

    except PlaywrightTimeout:
        result["status"] = "Timeout"
    except Exception as e:
        result["status"] = f"Error: {str(e)[:60]}"

    return result


# ── Read usernames from xlsx ─────────────────────────────────────────────────

def read_usernames(filepath: str, sheet_name: str = None, column: str = "A") -> list[str]:
    """Read usernames from a specific column of an xlsx sheet."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print(f"Error: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
            sys.exit(1)
        ws = wb[sheet_name]
    else:
        ws = wb.active

    usernames = []
    col_idx = openpyxl.utils.column_index_from_string(column)

    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
        val = row[0]
        if val and str(val).strip():
            clean = str(val).strip().lstrip("@")
            usernames.append(clean)

    wb.close()
    print(f"Loaded {len(usernames)} usernames from '{filepath}' (sheet: {sheet_name or 'active'})")
    return usernames


# ── Write results to xlsx ────────────────────────────────────────────────────

def write_results(results: list[dict], output_path: str):
    """Write scan results to a professionally formatted xlsx file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Twitter Scan Results"

    # ── Styles ────────────────────────────────────────────────────────────
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1DA1F2")  # Twitter blue
    header_align = Alignment(horizontal="center", vertical="center")

    data_font = Font(name="Arial", size=10)
    data_align_center = Alignment(horizontal="center", vertical="center")
    data_align_left = Alignment(horizontal="left", vertical="center")
    data_align_right = Alignment(horizontal="right", vertical="center")

    yes_fill = PatternFill("solid", fgColor="D4EDDA")   # green tint
    no_fill = PatternFill("solid", fgColor="F8D7DA")    # red tint
    error_fill = PatternFill("solid", fgColor="FFF3CD")  # yellow tint

    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    # ── Headers ───────────────────────────────────────────────────────────
    headers = ["#", "Username", "Blue Check Verified", "Followers", "Followers (Raw)", "Status"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # ── Data rows ─────────────────────────────────────────────────────────
    for i, r in enumerate(results, start=1):
        row_num = i + 1
        is_error = r["status"] not in ("OK", "Private Account")

        ws.cell(row=row_num, column=1, value=i).font = data_font
        ws.cell(row=row_num, column=1).alignment = data_align_center
        ws.cell(row=row_num, column=1).border = thin_border

        cell_user = ws.cell(row=row_num, column=2, value=f'@{r["username"]}')
        cell_user.font = Font(name="Arial", size=10, color="1DA1F2")
        cell_user.alignment = data_align_left
        cell_user.border = thin_border

        cell_verified = ws.cell(row=row_num, column=3, value=r["verified"])
        cell_verified.font = data_font
        cell_verified.alignment = data_align_center
        cell_verified.border = thin_border
        if r["verified"] == "Yes":
            cell_verified.fill = yes_fill
        elif not is_error:
            cell_verified.fill = no_fill

        cell_foll = ws.cell(row=row_num, column=4, value=r["followers"])
        cell_foll.font = data_font
        cell_foll.alignment = data_align_right
        cell_foll.border = thin_border

        cell_raw = ws.cell(row=row_num, column=5, value=r["followers_raw"] if r["followers_raw"] else "N/A")
        cell_raw.font = data_font
        cell_raw.alignment = data_align_right
        cell_raw.border = thin_border

        cell_status = ws.cell(row=row_num, column=6, value=r["status"])
        cell_status.font = data_font
        cell_status.alignment = data_align_center
        cell_status.border = thin_border
        if is_error:
            cell_status.fill = error_fill

    # ── Column widths ─────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 28

    # ── Freeze header row ─────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Auto-filter ───────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A1:F{len(results) + 1}"

    # ── Summary sheet ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    total = len(results)
    verified_count = sum(1 for r in results if r["verified"] == "Yes")
    not_verified = sum(1 for r in results if r["verified"] == "No" and r["status"] == "OK")
    errors = sum(1 for r in results if r["status"] not in ("OK", "Private Account"))
    private_count = sum(1 for r in results if r["status"] == "Private Account")

    summary_data = [
        ("Metric", "Value"),
        ("Total Usernames Scanned", total),
        ("Blue Check Verified (Yes)", verified_count),
        ("Not Verified (No)", not_verified),
        ("Private Accounts", private_count),
        ("Errors / Not Found", errors),
        ("Verification Rate", f"{verified_count / max(total, 1) * 100:.1f}%"),
    ]

    for row_idx, (label, value) in enumerate(summary_data, 1):
        cell_a = ws2.cell(row=row_idx, column=1, value=label)
        cell_b = ws2.cell(row=row_idx, column=2, value=value)
        cell_a.border = thin_border
        cell_b.border = thin_border
        if row_idx == 1:
            cell_a.font = header_font
            cell_a.fill = header_fill
            cell_b.font = header_font
            cell_b.fill = header_fill
        else:
            cell_a.font = Font(name="Arial", size=10, bold=True)
            cell_b.font = data_font
            cell_b.alignment = data_align_center

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 18

    wb.save(output_path)
    print(f"\nResults saved to: {output_path}")
    print(f"  Total: {total} | Verified: {verified_count} | Not Verified: {not_verified} | Errors: {errors}")


# ── Checkpoint helpers ───────────────────────────────────────────────────────

def save_checkpoint(results: list[dict], checkpoint_path: str):
    with open(checkpoint_path, "w") as f:
        json.dump(results, f, indent=2)

def load_checkpoint(checkpoint_path: str) -> list[dict]:
    if os.path.exists(checkpoint_path):
        with open(checkpoint_path, "r") as f:
            data = json.load(f)
        print(f"Resuming from checkpoint: {len(data)} usernames already scanned.")
        return data
    return []


# ── Main scan loop ───────────────────────────────────────────────────────────

def run_scan(
    input_file: str,
    sheet_name: str = None,
    column: str = "A",
    output_file: str = "twitter_results.xlsx",
    headless: bool = False,
    resume: bool = True,
):
    usernames = read_usernames(input_file, sheet_name, column)
    if not usernames:
        print("No usernames found. Exiting.")
        return

    checkpoint_path = output_file.replace(".xlsx", "_checkpoint.json")

    # Load checkpoint if resuming
    results = load_checkpoint(checkpoint_path) if resume else []
    already_done = {r["username"].lower() for r in results}
    remaining = [u for u in usernames if u.lower() not in already_done]

    if not remaining:
        print("All usernames already scanned. Writing final output...")
        write_results(results, output_file)
        return

    print(f"\nRemaining to scan: {len(remaining)}")
    print(f"Mode: {'Headless' if headless else 'Visible browser'}")
    print(f"Delay between requests: {MIN_DELAY}-{MAX_DELAY}s")
    print(f"Batch pause every {BATCH_PAUSE_EVERY} requests: {BATCH_PAUSE_SECONDS}s\n")

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=headless,
            args=["--disable-blink-features=AutomationControlled"],
        )
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/121.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1280, "height": 900},
            locale="en-US",
        )
        page = context.new_page()

        # Block unnecessary resources to speed things up
        def route_handler(route):
            if route.request.resource_type in ("image", "media", "font"):
                route.abort()
            else:
                route.continue_()

        page.route("**/*", route_handler)

        scan_count = 0
        for i, username in enumerate(remaining):
            scan_count += 1
            progress = len(results) + 1
            total = len(usernames)
            pct = progress / total * 100

            print(f"  [{progress}/{total}] ({pct:.0f}%) Scanning @{username}...", end=" ", flush=True)

            # Retry logic
            result = None
            for attempt in range(1, MAX_RETRIES + 1):
                wait_for_internet()
                result = scrape_profile(page, username)
                if result["status"] == "OK" or result["status"] == "Private Account":
                    break
                if result["status"] in ("Not Found / Suspended",):
                    break
                
                # Check for explicit network loss before consuming a retry
                if "net::ERR_INTERNET_DISCONNECTED" in result["status"] or result["status"] == "Timeout" or "Error" in result["status"]:
                    wait_for_internet()
                    
                if attempt < MAX_RETRIES:
                    print(f"(retry {attempt})...", end=" ", flush=True)
                    time.sleep(2)

            verified_label = "✓ Verified" if result["verified"] == "Yes" else "✗ Not Verified"
            print(f"{verified_label} | {result['followers']} followers | {result['status']}")

            results.append(result)

            # Checkpoint
            if scan_count % CHECKPOINT_EVERY == 0:
                save_checkpoint(results, checkpoint_path)
                print(f"    ↳ Checkpoint saved ({len(results)} scanned)")

            # Rate limiting
            if scan_count % BATCH_PAUSE_EVERY == 0 and i < len(remaining) - 1:
                print(f"\n  ⏸  Batch pause ({BATCH_PAUSE_SECONDS}s) to avoid rate limits...\n")
                time.sleep(BATCH_PAUSE_SECONDS)
            else:
                delay = random.uniform(MIN_DELAY, MAX_DELAY)
                time.sleep(delay)

        browser.close()

    # Final save
    save_checkpoint(results, checkpoint_path)
    write_results(results, output_file)

    # Clean up checkpoint
    if os.path.exists(checkpoint_path):
        os.remove(checkpoint_path)
        print("Checkpoint file cleaned up.")


# ── CLI ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Scan Twitter/X profiles from an xlsx list.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python twitter_scanner.py users.xlsx --sheet "Sheet1" --output results.xlsx
  python twitter_scanner.py users.xlsx --column B --headless
  python twitter_scanner.py users.xlsx --no-resume
        """,
    )
    parser.add_argument("input", help="Path to the .xlsx file containing usernames")
    parser.add_argument("--sheet", default=None, help="Sheet name to read from (default: active sheet)")
    parser.add_argument("--column", default="A", help="Column letter containing usernames (default: A)")
    parser.add_argument("--output", default="twitter_results.xlsx", help="Output .xlsx file path")
    parser.add_argument("--headless", action="store_true", help="Run browser in headless mode (invisible)")
    parser.add_argument("--no-resume", action="store_true", help="Start fresh, ignore any checkpoint")

    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"Error: File '{args.input}' not found.")
        sys.exit(1)

    run_scan(
        input_file=args.input,
        sheet_name=args.sheet,
        column=args.column,
        output_file=args.output,
        headless=args.headless,
        resume=not args.no_resume,
    )


if __name__ == "__main__":
    main()
